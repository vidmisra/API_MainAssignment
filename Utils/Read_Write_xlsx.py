import openpyxl
from openpyxl.styles import Alignment

path = f"/Users/vidmisra/PycharmProjects/API_MainAssignment/Utils/UserData.xlsx"


class XlsxReader:

    @staticmethod
    def write_data(email, response):
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        row_count = sheet.max_row
        c1 = sheet.cell(row=row_count + 1, column=1)
        c1.value = email
        c2 = sheet.cell(row=row_count + 1, column=2)
        c2.value = response.as_dict["data"]["description"]
        wb.save(path)
        return sheet.max_row

    def merge_cells(self, merge_range=19):
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        row_count = sheet.max_row
        col_count = sheet.max_column
        sheet.merge_cells('A{0}:A{1}'.format(row_count - merge_range, row_count))
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        wb.save(path)
